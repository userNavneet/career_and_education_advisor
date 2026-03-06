from datetime import datetime
from pathlib import Path
import os

from openpyxl import load_workbook

from database.db import get_db


XLSX_PATH = Path(
    os.getenv(
        "FAQ_XLSX_PATH",
        "/Users/cheshta/Documents/GitHub/career_and_education_advisor/data/faq_source.xlsx",
    )
)
FAQ_COLLECTION = os.getenv("FAQ_COLLECTION", "chatbot_faq")


def _normalize_header(value):
    return str(value or "").strip().lower()


def main():
    if not XLSX_PATH.exists():
        raise SystemExit(f"Missing FAQ file: {XLSX_PATH}")

    wb = load_workbook(XLSX_PATH, read_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        raise SystemExit("Excel file is empty")

    headers = [_normalize_header(h) for h in header_row]
    try:
        q_idx = headers.index("question")
    except ValueError:
        raise SystemExit(f"Missing 'question' column. Found: {headers}")

    # Handle both `answer` and `answer ` style headers.
    answer_index = -1
    for i, h in enumerate(headers):
        if h == "answer":
            answer_index = i
            break
    if answer_index < 0:
        raise SystemExit(f"Missing 'answer' column. Found: {headers}")

    db = get_db()
    collection = db[FAQ_COLLECTION]
    collection.create_index("question", unique=True)
    now = datetime.utcnow()

    inserted = 0
    updated = 0
    skipped = 0

    for row in rows:
        question = str(row[q_idx] or "").strip()
        answer = str(row[answer_index] or "").strip()

        if not question or not answer:
            skipped += 1
            continue

        result = collection.update_one(
            {"question": question},
            {
                "$set": {"answer": answer, "updated_at": now},
                "$setOnInsert": {"created_at": now},
            },
            upsert=True,
        )
        if result.upserted_id:
            inserted += 1
        else:
            updated += 1

    print(
        f"FAQ import complete in '{FAQ_COLLECTION}': "
        f"inserted={inserted}, updated={updated}, skipped={skipped}"
    )


if __name__ == "__main__":
    main()
